from openpyxl import load_workbook
from models import NursingHome

def parse_file(xlsx_path):
    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    nursing_homes = []
    for row in ws.iter_rows(min_row=2):  # Assuming row 1 is headers
        nursing_home = NursingHome(
            name=str(row[0].value) if row[0].value is not None else "",  # Ensure the value is a string
            address_line_1=str(row[1].value) if row[1].value is not None else "",
            suburb=str(row[2].value) if row[2].value is not None else "",
            state=str(row[3].value) if row[3].value is not None else "",
            post_code=str(row[4].value) if row[4].value is not None else "",
        )
        nursing_homes.append(nursing_home)
    return nursing_homes




